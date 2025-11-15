"""
Excel file reading module with chunked processing support
"""
import pandas as pd
import logging
from typing import Generator, Tuple, Optional
from datetime import datetime
import sys
import os

# Add parent directory to path for imports
sys.path.append(os.path.dirname(os.path.dirname(os.path.dirname(__file__))))
import config

logger = logging.getLogger(__name__)


def extract_start_date_from_summary(file_path: str) -> Optional[str]:
    """
    Extract start date from the Summary sheet and return it as "YYYY-MM" format.

    Looks for a cell containing "Start date:" and extracts the date value from the adjacent cell.

    Args:
        file_path: Path to the Excel file

    Returns:
        Date string in "YYYY-MM" format, or None if not found
    """
    try:
        from openpyxl import load_workbook

        wb = load_workbook(file_path, read_only=True, data_only=True)

        # Try to get Summary sheet
        if 'Summary' not in wb.sheetnames:
            logger.warning("Summary sheet not found in Excel file")
            wb.close()
            return None

        ws = wb['Summary']

        # Search for "Start date:" in the sheet
        for row in ws.iter_rows(values_only=True):
            for idx, cell_value in enumerate(row):
                if cell_value and isinstance(cell_value, str) and 'Start date:' in cell_value:
                    # Found "Start date:", now get the next cell value
                    if idx + 1 < len(row):
                        date_value = row[idx + 1]
                        logger.info(f"Found start date cell value: {date_value} (type: {type(date_value)})")

                        # Try to parse the date
                        if date_value:
                            if isinstance(date_value, datetime):
                                # Already a datetime object
                                date_str = date_value.strftime("%Y-%m")
                                logger.info(f"Extracted date from Summary sheet: {date_str}")
                                wb.close()
                                return date_str
                            elif isinstance(date_value, str):
                                # Try to parse string date
                                try:
                                    parsed_date = pd.to_datetime(date_value)
                                    date_str = parsed_date.strftime("%Y-%m")
                                    logger.info(f"Extracted date from Summary sheet: {date_str}")
                                    wb.close()
                                    return date_str
                                except:
                                    logger.warning(f"Could not parse date string: {date_value}")

        logger.warning("Could not find 'Start date:' in Summary sheet")
        wb.close()
        return None

    except Exception as e:
        logger.error(f"Error extracting start date from Summary sheet: {e}")
        return None


def get_total_rows(file_path: str, sheet_name: str = 'Data') -> int:
    """
    Get the total number of rows in an Excel file without loading the entire file.

    Args:
        file_path: Path to the Excel file
        sheet_name: Name of the sheet to read (default: 'Data')

    Returns:
        Total number of rows in the file
    """
    try:
        # Use openpyxl to get row count efficiently
        from openpyxl import load_workbook
        wb = load_workbook(file_path, read_only=True)

        # Get the specified sheet
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            logger.warning(f"Sheet '{sheet_name}' not found, using active sheet")
            ws = wb.active

        total_rows = ws.max_row - 1  # Subtract 1 for header row
        wb.close()
        return total_rows
    except Exception as e:
        logger.warning(f"Could not get exact row count, using pandas: {e}")
        # Fallback to pandas if openpyxl fails
        df = pd.read_excel(file_path, sheet_name=sheet_name, nrows=0)
        # This is less accurate but works as a fallback
        return 0


def read_excel_in_chunks(
    file_path: str,
    chunk_size: int = None,
    sheet_name: str = 'Data'
) -> Generator[Tuple[pd.DataFrame, int], None, None]:
    """
    Read Excel file in chunks to optimize memory usage for large files.

    This function yields chunks of data as DataFrames along with the total row count,
    allowing for progress tracking and memory-efficient processing.

    Args:
        file_path: Path to the Excel file
        chunk_size: Number of rows per chunk (default from config)
        sheet_name: Name of the sheet to read (default: 'Data')

    Yields:
        Tuple of (chunk_dataframe, total_rows)

    Example:
        >>> for chunk, total in read_excel_in_chunks('large_file.xlsx'):
        ...     process_chunk(chunk)
        ...     print(f"Processed {len(chunk)} rows out of {total}")
    """
    if chunk_size is None:
        chunk_size = config.CHUNK_SIZE

    logger.info(f"Reading Excel file: {file_path} from sheet '{sheet_name}' with chunk size: {chunk_size}")

    try:
        from openpyxl import load_workbook

        # Load workbook in read-only mode for memory efficiency
        wb = load_workbook(file_path, read_only=True, data_only=True)

        # Get the specified sheet
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            logger.info(f"Successfully found sheet '{sheet_name}'")
        else:
            logger.warning(f"Sheet '{sheet_name}' not found, using active sheet")
            ws = wb.active

        # Get total rows
        total_rows = ws.max_row - 1  # Exclude header

        # Get header row
        header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

        # Read data in chunks
        current_row = 2  # Start from row 2 (after header)

        while current_row <= ws.max_row:
            rows_data = []

            # Read chunk_size rows
            for row in ws.iter_rows(min_row=current_row, max_row=current_row + chunk_size - 1, values_only=True):
                rows_data.append(row)

            if not rows_data:
                break

            # Create DataFrame from chunk
            chunk_df = pd.DataFrame(rows_data, columns=header)

            logger.debug(f"Yielding chunk with {len(chunk_df)} rows")
            yield chunk_df, total_rows

            current_row += chunk_size

        wb.close()

    except Exception as e:
        logger.error(f"Error reading Excel file: {e}")
        raise


def read_excel_file(file_path: str, sheet_name: str = 'Data') -> pd.DataFrame:
    """
    Read entire Excel file into a DataFrame.

    Use this for smaller files. For large files, use read_excel_in_chunks instead.

    Args:
        file_path: Path to the Excel file
        sheet_name: Name of the sheet to read (default: 'Data')

    Returns:
        DataFrame containing the Excel data
    """
    logger.info(f"Reading entire Excel file: {file_path} from sheet '{sheet_name}'")

    try:
        df = pd.read_excel(file_path, sheet_name=sheet_name, engine='openpyxl')
        logger.info(f"Successfully read {len(df)} rows")
        return df
    except Exception as e:
        logger.error(f"Error reading Excel file: {e}")
        raise
